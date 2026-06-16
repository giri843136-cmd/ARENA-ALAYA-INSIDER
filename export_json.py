#!/usr/bin/env python3
"""Export products.json from the latest pipeline output directory."""
import openpyxl
import json
import os

PROJECT_DIR = r"C:\Users\Admin\Downloads\Alaya Insider Product"
OUTPUT_BASE = os.path.join(PROJECT_DIR, "output")

def find_latest_output():
    """Find the most recent output directory."""
    dirs = [d for d in os.listdir(OUTPUT_BASE) if d.startswith("alaya_enrichment_")]
    if not dirs:
        raise FileNotFoundError("No output directories found.")
    latest = sorted(dirs)[-1]
    return os.path.join(OUTPUT_BASE, latest)

def find_xlsx(output_dir):
    """Find the products_import.xlsx in the given directory."""
    path = os.path.join(output_dir, "products_import.xlsx")
    if os.path.exists(path):
        return path
    raise FileNotFoundError(f"products_import.xlsx not found in {output_dir}")

def main():
    output_dir = find_latest_output()
    xlsx_path = find_xlsx(output_dir)
    print(f"Reading: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Products"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    products = []
    for row_idx in range(2, ws.max_row + 1):
        p = {}
        for col_idx, h in enumerate(headers, 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if h in ("Price", "Rating", "Review Count"):
                try:
                    val = float(val) if val else 0
                except:
                    val = 0
                if h == "Review Count":
                    val = int(val)
            p[h] = str(val) if not isinstance(val, (int, float)) else val
        products.append(p)

    out = [
        {
            "name": p.get("Product Name", ""),
            "brand": p.get("Brand", ""),
            "price": p.get("Price", 0),
            "rating": p.get("Rating", 0),
            "reviews": int(p.get("Review Count", 0)),
            "asin": str(p.get("ASIN", "")).replace(".0", ""),
            "link": p.get("Affiliate Link", ""),
            "category": p.get("Category", ""),
            "subcategory": p.get("Subcategory", ""),
            "merchant": p.get("Merchant Name", ""),
            "seo_title": p.get("SEO Title", ""),
            "short_desc": str(p.get("Short Description", ""))[:200],
            "features": str(p.get("Features", "")).split(" | ")[:5],
        }
        for p in products
    ]

    json_path = os.path.join(PROJECT_DIR, "products.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    size_kb = os.path.getsize(json_path) / 1024
    print(f"products.json generated: {len(out)} products ({size_kb:.0f} KB)")
    print(f"From: {output_dir}")

if __name__ == "__main__":
    main()
