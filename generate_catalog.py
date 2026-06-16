#!/usr/bin/env python3
"""Generate a static HTML catalog page from the latest products_import.xlsx."""
import json
import openpyxl
import os
from datetime import datetime

PROJECT_DIR = r"C:\Users\Admin\Downloads\Alaya Insider Product"
OUTPUT_BASE = os.path.join(PROJECT_DIR, "output")

# Auto-detect latest output directory
dirs = [d for d in os.listdir(OUTPUT_BASE) if d.startswith("alaya_enrichment_")]
if not dirs:
    raise FileNotFoundError("No output directories found.")
latest_dir = sorted(dirs)[-1]
EXCEL_PATH = os.path.join(OUTPUT_BASE, latest_dir, "products_import.xlsx")
OUTPUT_PATH = os.path.join(PROJECT_DIR, "catalog.html")

# Load real Amazon product images
PRODUCT_IMAGES_PATH = os.path.join(PROJECT_DIR, "product_images.json")
product_images = {}
real_image_count = 0
if os.path.exists(PRODUCT_IMAGES_PATH):
    with open(PRODUCT_IMAGES_PATH, 'r', encoding='utf-8') as f:
        product_images = json.load(f)
    real_image_count = sum(1 for v in product_images.values() if v)
    print(f"Loaded {real_image_count} real Amazon product images")

print(f"Using latest output: {latest_dir}")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Products']
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

products = []
for row_idx in range(2, ws.max_row + 1):
    p = {}
    for col_idx, h in enumerate(headers, 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if h in ('Price', 'Rating', 'Review Count'):
            try:
                val = float(val) if val else 0
            except:
                val = 0
        p[h] = str(val) if not isinstance(val, (int, float)) else val
    products.append(p)

total = len(products)
avg_price = sum(float(p.get('Price', 0)) for p in products) / total if total else 0
avg_rating = sum(float(p.get('Rating', 0)) for p in products) / total if total else 0
total_reviews = sum(int(p.get('Review Count', 0)) for p in products)

img_variants = [
    'https://images.unsplash.com/photo-1555041469-a586c61ea9bc?w=600&q=80',
    'https://images.unsplash.com/photo-1544716278-ca5e3f4abd8c?w=600&q=80',
    'https://images.unsplash.com/photo-1484101403633-562f891dc89a?w=600&q=80',
    'https://images.unsplash.com/photo-1506439773649-6e0eb8cfb237?w=600&q=80',
    'https://images.unsplash.com/photo-1544457070-4cd773b4d71e?w=600&q=80',
    'https://images.unsplash.com/photo-1567016432779-094069958ea5?w=600&q=80',
    'https://images.unsplash.com/photo-1586023492125-27b2c045efd7?w=600&q=80',
    'https://images.unsplash.com/photo-1522708323590-d24dbb6b0267?w=600&q=80',
    'https://images.unsplash.com/photo-1513506003901-1e6a229e2d15?w=600&q=80',
    'https://images.unsplash.com/photo-1493663284031-b7e3aefcae8e?w=600&q=80',
    'https://images.unsplash.com/photo-1540575467063-178a50c2df87?w=600&q=80',
    'https://images.unsplash.com/photo-1449247709967-d4461a6a6103?w=600&q=80',
]

# Collect unique categories
all_categories = sorted(set(
    p.get('Category', '').split(' >')[0] if '>' in p.get('Category', '') else p.get('Category', '')
    for p in products
))
category_options = ''.join(f'<option value="{c.replace(chr(34), "&quot;")}">{c}</option>' for c in all_categories)

card_template = '''\
    <div class="product-card {verified}" data-name="{data_name}" data-category="{category}" data-price="{price}" data-rating="{rating}" data-brand="{data_brand}">
      <div class="product-image">
        <img src="{img}" alt="{name}" loading="lazy">
        <div class="product-tags">
          <span class="tag category-tag">{category}</span>
          <span class="tag subcat-tag">{subcat}</span>
        </div>
      </div>
      <div class="product-info">
        <div class="brand">{brand}</div>
        <h3 class="product-name">{name_display}</h3>
        <div class="rating">
          <span class="stars">{stars}</span>
          <span class="review-count">({reviews:,})</span>
        </div>
        <p class="desc">{desc}</p>
        <div class="specs">
          <span class="spec"><strong>Color:</strong> {color}</span>
          <span class="spec"><strong>Material:</strong> {material}</span>
          <span class="spec"><strong>ASIN:</strong> {asin}</span>
        </div>
        <ul class="features">{features}</ul>
        <div class="price-row">
          <span class="price">${price:,.2f}</span>
          <span class="merchant">{merchant}</span>
        </div>
        <a href="{link}" target="_blank" class="shop-btn">Shop Now &rarr;</a>
      </div>
    </div>'''

cards_html = ""
for i, p in enumerate(products):
    name = p.get('Product Name', '')
    brand = p.get('Brand', '')
    price_val = float(p.get('Price', 0))
    rating_val = float(p.get('Rating', 0))
    reviews = int(p.get('Review Count', 0))
    asin = str(p.get('ASIN', '')).replace('.0', '')[:14]
    link = p.get('Affiliate Link', '')
    cat = p.get('Category', '').split(' >')[0] if '>' in p.get('Category', '') else p.get('Category', '')
    subcat = p.get('Subcategory', '')
    merchant = p.get('Merchant Name', '')
    short_desc = str(p.get('Short Description', ''))[:150]
    features_raw = str(p.get('Features', '')).split(' | ')[:4]
    features_list = ''.join(f'<li>{f}</li>' for f in features_raw if f)
    color = p.get('Color', '')
    material = str(p.get('Material', '')).split(',')[0].strip()
    # Use real Amazon image if available for this ASIN, otherwise fall back to Unsplash
    real_img = product_images.get(asin, '') if asin else ''
    if real_img:
        # Use the base image URL (remove size modifiers for cleaner display)
        img = real_img
    else:
        img = img_variants[i % len(img_variants)]
    verified = 'verified' if asin.startswith('B0') and len(asin.replace('.0', '')) >= 10 else ''
    name_display = name[:52] + '...' if len(name) > 52 else name

    full_stars = int(rating_val)
    empty_stars = 5 - full_stars
    stars = '★' * full_stars + '☆' * empty_stars

    card = card_template.format(
        verified=verified, img=img, name=name.replace('"', '&quot;'),
        data_name=name.replace('"', '&quot;').lower(),
        data_brand=brand.lower().replace('"', ''),
        category=cat, subcat=subcat, brand=brand,
        name_display=name_display.replace('"', '&quot;'),
        stars=stars, reviews=reviews, desc=short_desc.replace('"', '&quot;'),
        color=color, material=material, asin=asin,
        features=features_list, price=price_val, merchant=merchant,
        rating=rating_val,
        link=link.replace('"', '&quot;')
    )
    cards_html += card

html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ALAYA INSIDER - Curated Product Catalog</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Playfair+Display:wght@600;700&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'DM Sans',sans-serif; background:#faf9f8; color:#1a1a2e; }}
.header {{
  background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
  color:white; padding:50px 20px 40px; text-align:center; position:relative; overflow:hidden;
}}
.header::before {{
  content:''; position:absolute; top:-50%; left:-50%; width:200%; height:200%;
  background:radial-gradient(circle,rgba(226,167,111,0.08) 0%,transparent 60%);
}}
.header h1 {{ font-family:'Playfair Display',serif; font-size:2.8em; letter-spacing:3px; position:relative; }}
.header h1 span {{ color:#e2a76f; }}
.header p {{ margin-top:10px; opacity:0.8; font-size:1.05em; position:relative; }}
.header .stats {{ display:flex; justify-content:center; gap:40px; margin-top:25px; flex-wrap:wrap; position:relative; }}
.header .stat {{ text-align:center; }}
.header .stat-value {{ font-size:1.8em; font-weight:700; color:#e2a76f; }}
.header .stat-label {{ font-size:0.78em; opacity:0.7; text-transform:uppercase; letter-spacing:1.5px; }}

.toolbar {{
  background:white; padding:16px 20px; display:flex; gap:12px; align-items:center;
  flex-wrap:wrap; border-bottom:1px solid #e8e6e3; position:sticky; top:0; z-index:10;
  box-shadow:0 2px 8px rgba(0,0,0,0.04);
}}
.toolbar .search-box {{
  flex:1; min-width:200px; padding:10px 16px; border:1px solid #ddd; border-radius:8px;
  font-family:'DM Sans',sans-serif; font-size:0.9em; outline:none; transition:border-color 0.2s;
}}
.toolbar .search-box:focus {{ border-color:#e2a76f; box-shadow:0 0 0 3px rgba(226,167,111,0.15); }}
.toolbar select {{
  padding:10px 14px; border:1px solid #ddd; border-radius:8px; font-family:'DM Sans',sans-serif;
  font-size:0.85em; outline:none; background:white; cursor:pointer; transition:border-color 0.2s;
}}
.toolbar select:focus {{ border-color:#e2a76f; }}
.toolbar .result-count {{ font-size:0.85em; color:#888; white-space:nowrap; }}
.toolbar .clear-btn {{
  padding:8px 16px; border:1px solid #ddd; border-radius:8px; background:white;
  font-family:'DM Sans',sans-serif; font-size:0.85em; cursor:pointer; transition:all 0.2s;
  color:#666;
}}
.toolbar .clear-btn:hover {{ background:#f5f4f2; border-color:#ccc; }}

.container {{ max-width:1400px; margin:0 auto; padding:20px; }}
.grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(320px,1fr)); gap:20px; }}
.product-card {{
  background:white; border-radius:14px; overflow:hidden;
  box-shadow:0 2px 20px rgba(0,0,0,0.06); transition:all 0.4s cubic-bezier(0.25,0.46,0.45,0.94);
  position:relative; animation:fadeIn 0.3s ease;
}}
@keyframes fadeIn {{ from {{ opacity:0; transform:translateY(10px); }} to {{ opacity:1; transform:translateY(0); }} }}
.product-card.hidden {{ display:none; }}
.product-card:hover {{ transform:translateY(-6px); box-shadow:0 12px 40px rgba(0,0,0,0.12); }}
.product-card.verified::before {{
  content:'VERIFIED'; position:absolute; top:12px; right:12px;
  background:#e2a76f; color:#1a1a2e; font-size:9px; font-weight:700;
  padding:4px 10px; border-radius:20px; letter-spacing:1px; z-index:2;
}}
.product-image {{ position:relative; height:220px; overflow:hidden; background:#f0eeeb; }}
.product-image img {{ width:100%; height:100%; object-fit:cover; transition:transform 0.6s ease; }}
.product-card:hover .product-image img {{ transform:scale(1.06); }}
.product-tags {{ position:absolute; bottom:10px; left:10px; display:flex; gap:5px; }}
.tag {{ font-size:9px; font-weight:600; padding:3px 10px; border-radius:20px; text-transform:uppercase; letter-spacing:0.5px; }}
.category-tag {{ background:rgba(26,26,46,0.85); color:white; }}
.subcat-tag {{ background:rgba(226,167,111,0.9); color:#1a1a2e; }}
.product-info {{ padding:18px; }}
.brand {{ font-size:10px; font-weight:600; text-transform:uppercase; letter-spacing:1.5px; color:#e2a76f; margin-bottom:3px; }}
.product-name {{ font-family:'Playfair Display',serif; font-size:1.1em; line-height:1.3; margin-bottom:8px; color:#1a1a2e; }}
.rating {{ display:flex; align-items:center; gap:6px; margin-bottom:10px; }}
.stars {{ color:#e2a76f; font-size:0.9em; letter-spacing:1.5px; }}
.review-count {{ font-size:0.78em; color:#999; }}
.desc {{ font-size:0.82em; color:#666; line-height:1.5; margin-bottom:10px; }}
.specs {{ display:flex; flex-wrap:wrap; gap:6px; margin-bottom:10px; }}
.spec {{ font-size:0.75em; color:#888; background:#f5f4f2; padding:2px 8px; border-radius:10px; }}
.features {{ list-style:none; margin-bottom:12px; }}
.features li {{ font-size:0.78em; color:#555; padding:2px 0 2px 16px; position:relative; }}
.features li::before {{ content:'\\2713'; position:absolute; left:0; color:#e2a76f; font-weight:700; }}
.price-row {{
  display:flex; justify-content:space-between; align-items:center;
  margin-bottom:12px; padding-top:12px; border-top:1px solid #f0eeeb;
}}
.price {{ font-size:1.3em; font-weight:700; color:#1a1a2e; }}
.merchant {{ font-size:0.78em; color:#aaa; }}
.shop-btn {{
  display:block; background:linear-gradient(135deg,#1a1a2e,#16213e);
  color:white; text-align:center; padding:11px; border-radius:8px;
  text-decoration:none; font-weight:600; font-size:0.85em;
  transition:all 0.3s ease;
}}
.shop-btn:hover {{
  background:linear-gradient(135deg,#16213e,#0f3460);
  transform:translateY(-1px); box-shadow:0 6px 20px rgba(15,52,96,0.3);
}}
.footer {{ text-align:center; padding:30px 20px; color:#aaa; font-size:0.82em; border-top:1px solid #e8e6e3; margin-top:30px; }}
.footer a {{ color:#e2a76f; text-decoration:none; font-weight:500; }}
.footer .small {{ font-size:0.82em; opacity:0.7; margin-top:4px; display:block; }}
@media (max-width:768px) {{
  .grid {{ grid-template-columns:1fr; }}
  .header h1 {{ font-size:2em; }}
  .header .stats {{ gap:20px; }}
  .header .stat-value {{ font-size:1.4em; }}
  .toolbar {{ flex-direction:column; }}
  .toolbar .search-box {{ width:100%; }}
  .toolbar select {{ width:100%; }}
}}
</style>
</head>
<body>

<div class="header">
  <h1>ALAYA <span>INSIDER</span></h1>
  <p>Curated Product Catalog &mdash; {total} premium hand-picked products</p>
  <div class="stats">
    <div class="stat"><div class="stat-value">{total}</div><div class="stat-label">Products</div></div>
    <div class="stat"><div class="stat-value">${avg_price:,.0f}</div><div class="stat-label">Avg Price</div></div>
    <div class="stat"><div class="stat-value">{avg_rating:.1f}</div><div class="stat-label">Avg Rating</div></div>
    <div class="stat"><div class="stat-value">{total_reviews:,}</div><div class="stat-label">Total Reviews</div></div>
  </div>
</div>

<div class="toolbar">
  <input type="text" class="search-box" id="searchInput" placeholder="Search products by name, brand, or category..." oninput="debouncedFilter()">
  <select id="categoryFilter" onchange="filterProducts()">
    <option value="all">All Categories</option>
    {category_options}
  </select>
  <select id="sortSelect" onchange="filterProducts()">
    <option value="default">Sort: Default</option>
    <option value="price-asc">Price: Low to High</option>
    <option value="price-desc">Price: High to Low</option>
    <option value="rating-desc">Rating: High to Low</option>
    <option value="name-asc">Name: A-Z</option>
  </select>
  <span class="result-count" id="resultCount">Showing {total} products</span>
  <button class="clear-btn" onclick="clearFilters()">Clear</button>
</div>

<div class="container">
  <div class="grid" id="productGrid">
    {cards_html}
  </div>
</div>

<div class="footer">
  Curated by <a href="#">ALAYA INSIDER</a> &mdash; Product Intelligence System
  <span class="small">Affiliate links &bull; Prices subject to change &bull; Data as of {datetime.now().strftime("%B %d, %Y")}</span>
</div>

<script>
function filterProducts() {{
  const search = document.getElementById('searchInput').value.toLowerCase().trim();
  const category = document.getElementById('categoryFilter').value;
  const sort = document.getElementById('sortSelect').value;
  const grid = document.getElementById('productGrid');
  const cards = grid.querySelectorAll('.product-card');
  let visible = 0;

  // Build array of visible cards
  const visibleCards = [];
  cards.forEach(card => {{
    const name = card.getAttribute('data-name') || '';
    const brand = card.getAttribute('data-brand') || '';
    const cat = card.getAttribute('data-category') || '';
    const price = parseFloat(card.getAttribute('data-price')) || 0;
    const rating = parseFloat(card.getAttribute('data-rating')) || 0;

    let show = true;
    if (search && !name.includes(search) && !brand.includes(search)) show = false;
    if (category !== 'all' && cat !== category) show = false;

    card.classList.toggle('hidden', !show);
    if (show) {{
      visible++;
      visibleCards.push({{ el: card, price, rating, name }});
    }}
  }});

  // Sort
  if (sort === 'price-asc') {{
    visibleCards.sort((a, b) => a.price - b.price);
  }} else if (sort === 'price-desc') {{
    visibleCards.sort((a, b) => b.price - a.price);
  }} else if (sort === 'rating-desc') {{
    visibleCards.sort((a, b) => b.rating - a.rating);
  }} else if (sort === 'name-asc') {{
    visibleCards.sort((a, b) => a.name.localeCompare(b.name));
  }}

  // Re-append sorted cards
  if (sort !== 'default' && visibleCards.length > 0) {{
    visibleCards.forEach(item => grid.appendChild(item.el));
  }}

  document.getElementById('resultCount').textContent = 'Showing ' + visible + ' of ' + cards.length + ' products';
}}

let debounceTimer;
function debouncedFilter() {{
  clearTimeout(debounceTimer);
  debounceTimer = setTimeout(filterProducts, 150);
}}

function clearFilters() {{
  document.getElementById('searchInput').value = '';
  document.getElementById('categoryFilter').value = 'all';
  document.getElementById('sortSelect').value = 'default';
  filterProducts();
}}
</script>

</body>
</html>'''

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Catalog generated: {OUTPUT_PATH}")
print(f"Products: {total}")
print(f"Avg Price: ${avg_price:,.2f}")
print(f"Avg Rating: {avg_rating:.1f}")
print(f"Total Reviews: {total_reviews:,}")
print(f"Categories: {len(all_categories)}")
